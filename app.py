"""
MOOVLY FLEET OPTIMIZER
Flask Backend
"""

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
import os
import json
import uuid
from datetime import datetime
from moovly_system import (
    charger_destinations_excel, charger_employes_excel,
    geocoder_lieu, generer_suggestions, calculer_tarif
)

app = Flask(__name__)


# DATABASE CONFIG
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///moovly_fleet.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


# MODELS
class Employee(db.Model):
    __tablename__ = 'employees'
    id        = db.Column(db.Integer, primary_key=True)
    emp_id    = db.Column(db.String(50))
    _id       = db.Column(db.String(50), unique=True, nullable=False, index=True)
    nom       = db.Column(db.String(200))
    residence = db.Column(db.String(500))
    lat       = db.Column(db.Float, nullable=True)
    lng       = db.Column(db.Float, nullable=True)

    def to_dict(self):
        return {
            'id': self.emp_id,
            'nom': self.nom,
            'residence': self.residence,
            'lat': self.lat,
            'lng': self.lng,
            '_id': self._id
        }

    @classmethod
    def from_dict(cls, d):
        return cls(
            emp_id    = d.get('id'),
            _id       = d.get('_id'),
            nom       = d.get('nom'),
            residence = d.get('residence'),
            lat       = d.get('lat'),
            lng       = d.get('lng')
        )


class Destination(db.Model):
    __tablename__ = 'destinations'
    id      = db.Column(db.Integer, primary_key=True)
    dest_id = db.Column(db.String(50))
    _id     = db.Column(db.String(50), unique=True, nullable=False, index=True)
    nom     = db.Column(db.String(200))
    lat     = db.Column(db.Float, nullable=True)
    lng     = db.Column(db.Float, nullable=True)

    def to_dict(self):
        return {
            'id': self.dest_id,
            'nom': self.nom,
            'lat': self.lat,
            'lng': self.lng,
            '_id': self._id
        }

    @classmethod
    def from_dict(cls, d):
        return cls(
            dest_id = d.get('id'),
            _id     = d.get('_id'),
            nom     = d.get('nom'),
            lat     = d.get('lat'),
            lng     = d.get('lng')
        )


class OptimizationHistory(db.Model):
    __tablename__ = 'optimization_history'
    id              = db.Column(db.Integer, primary_key=True)
    snapshot_id     = db.Column(db.String(50), default=lambda: str(uuid.uuid4())[:8])
    date            = db.Column(db.DateTime, default=datetime.now)
    nb_employes     = db.Column(db.Integer, default=0)
    nb_vehicules    = db.Column(db.Integer, default=0)
    capacite        = db.Column(db.String(20), default='')
    distance_km     = db.Column(db.Float, default=0)
    duree_min       = db.Column(db.Float, default=0)
    tarif_tnd       = db.Column(db.Float, default=0)
    co2_kg          = db.Column(db.Float, default=0)
    co2_saved_kg    = db.Column(db.Float, default=0)
    cost_saved_tnd  = db.Column(db.Float, default=0)
    methode         = db.Column(db.String(100), default='')
    destination_nom = db.Column(db.String(200), default='')

    def to_dict(self):
        return {
            "id": self.snapshot_id,
            "date": self.date.isoformat() if self.date else None,
            "nb_employes": self.nb_employes,
            "nb_vehicules": self.nb_vehicules,
            "capacite": self.capacite,
            "distance_km": self.distance_km,
            "duree_min": self.duree_min,
            "tarif_tnd": self.tarif_tnd,
            "co2_kg": self.co2_kg,
            "co2_saved_kg": self.co2_saved_kg,
            "cost_saved_tnd": self.cost_saved_tnd,
            "methode": self.methode,
            "destination_nom": self.destination_nom
        }


# Auto-create tables on startup
with app.app_context():
    db.create_all()



# HELPERS
def _normaliser_capacite(capacite):
    if capacite is None:
        return 3
    if isinstance(capacite, str):
        text = capacite.strip().lower()
        if text == 'ia':
            return 'ia'
        try:
            return int(text)
        except ValueError:
            return 3
    if isinstance(capacite, (list, tuple)):
        return [int(c) for c in capacite]
    return int(capacite)



# PAGES
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')



# UPLOAD EMPLOYÉS
@app.route('/api/upload_employes', methods=['POST'])
def upload_employes():
    file = request.files.get('file')
    if not file:
        return jsonify({'status': 'error', 'message': 'Aucun fichier'}), 400
    filepath = os.path.join('uploads', file.filename)
    os.makedirs('uploads', exist_ok=True)
    file.save(filepath)
    try:
        employes = charger_employes_excel(filepath)
    except ValueError as ve:
        return jsonify({'status': 'error', 'message': str(ve)}), 400

    if not employes:
        return jsonify({'status': 'error', 'message': 'Aucun employé trouvé'}), 400

    geocode_failures = []
    for emp in employes:
        if emp['lat'] is None and emp.get('residence'):
            coords = geocoder_lieu(emp['residence'])
            if coords:
                emp['lat'], emp['lng'] = coords[0], coords[1]
            else:
                geocode_failures.append(emp['residence'])
        elif emp['lat'] is None:
            geocode_failures.append(emp.get('nom', 'Inconnu'))

    # Persist to SQLite
    try:
        Employee.query.delete()
        db.session.commit()
        for emp in employes:
            db.session.add(Employee.from_dict(emp))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({
        'status': 'success',
        'employes': employes,
        'count': len(employes),
        'geocode_failures': geocode_failures
    })



# DESTINATIONS
@app.route('/api/load_destinations')
def load_destinations():
    Destination.query.delete()
    db.session.commit()
    return jsonify({'status': 'success', 'destinations': [], 'count': 0})



# ADD MANUAL POINT
@app.route('/api/add_manual', methods=['POST'])
def add_manual():
    data = request.json
    data_type, item = data.get('type'), data.get('item')
    if data_type not in ['employes', 'destinations'] or not item:
        return jsonify({'status': 'error', 'message': 'Type ou item invalide'}), 400
    item['_id'] = f"manual_{os.urandom(4).hex()}"

    try:
        if data_type == 'destinations':
            Destination.query.delete()
            db.session.commit()
            db.session.add(Destination.from_dict(item))
        else:
            db.session.add(Employee.from_dict(item))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', '_id': item['_id']})



# CLEAR STORE (RÉINITIALISATION)
@app.route('/api/clear_store', methods=['POST'])
def clear_store():
    try:
        Employee.query.delete()
        Destination.query.delete()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    return jsonify({'status': 'success', 'message': 'Store vidé'})



# UPDATE LOCATION
@app.route('/api/update_location', methods=['POST'])
def update_location():
    data = request.json
    data_type, item_id = data.get('type'), data.get('id')
    lat, lng = data.get('lat'), data.get('lng')
    if data_type not in ['employes', 'destinations'] or not item_id or lat is None or lng is None:
        return jsonify({'status': 'error', 'message': 'Données invalides'}), 400

    if data_type == 'employes':
        entity = Employee.query.filter_by(_id=item_id).first()
    else:
        entity = Destination.query.filter_by(_id=item_id).first()

    if not entity:
        return jsonify({'status': 'error', 'message': 'Entité non trouvée'}), 404

    try:
        entity.lat = float(lat)
        entity.lng = float(lng)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', 'item': entity.to_dict()})



# ROUTE GEOMETRY
@app.route('/api/route_geometry', methods=['POST'])
def route_geometry():
    data = request.json
    waypoints = data.get('waypoints', [])
    use_alternative = data.get('use_alternative', False)

    if len(waypoints) < 2:
        return jsonify({'status': 'error', 'message': 'Min 2 waypoints'}), 400

    try:
        import requests as req
        import math

        def fetch_osrm(wps, alternatives=False):
            coords_str = ';'.join([f"{wp[1]},{wp[0]}" for wp in wps])
            url = f"http://router.project-osrm.org/route/v1/driving/{coords_str}"
            params = {'overview': 'false', 'steps': 'true', 'geometries': 'geojson'}
            if alternatives:
                params['alternatives'] = 'true'
            resp = req.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                rdata = resp.json()
                if rdata['code'] == 'Ok':
                    return rdata['routes']
            return None

        routes = fetch_osrm(waypoints, alternatives=use_alternative)

        if not routes:
            return jsonify({'status': 'fallback', 'latlngs': waypoints})

        route = None
        if use_alternative:
            if len(routes) > 1:
                route = routes[1]
            else:
                lat1, lng1 = waypoints[0]
                lat2, lng2 = waypoints[1]
                mid_lat = (lat1 + lat2) / 2.0
                mid_lng = (lng1 + lng2) / 2.0
                dx = lng2 - lng1
                dy = lat2 - lat1
                length = math.hypot(dx, dy)
                if length > 0.001:
                    nx = -dy / length
                    ny = dx / length
                    offset_lat = mid_lat + ny * 0.015
                    offset_lng = mid_lng + nx * 0.015
                    alt_waypoints = [waypoints[0], [offset_lat, offset_lng]] + waypoints[1:]
                    alt_routes = fetch_osrm(alt_waypoints, alternatives=False)
                    if alt_routes:
                        route = alt_routes[0]

                if not route:
                    return jsonify({'status': 'no_alternative', 'latlngs': waypoints})
        else:
            route = routes[0]

        latlngs = []
        for idx, leg in enumerate(route['legs']):
            for step in leg['steps']:
                for c in step['geometry']['coordinates']:
                    latlngs.append([c[1], c[0]])

        return jsonify({
            'status': 'success',
            'latlngs': latlngs,
            'distance_km': route['distance'] / 1000,
            'duration_min': route['duration'] / 60
        })
    except Exception as e:
        print("Erreur OSRM:", e)
        return jsonify({'status': 'fallback', 'latlngs': waypoints})



# OPTIMIZE
@app.route('/api/optimize', methods=['POST'])
def optimize():
    data = request.json or {}
    selected_emp_ids = data.get('selected_employes_ids', [])
    dest_id = data.get('destination_id')
    capacite = _normaliser_capacite(data.get('capacite', 3))

    poids_raw = data.get('poids', {})
    poids = {
        'distance': float(poids_raw.get('distance', 0.33)),
        'cout':     float(poids_raw.get('cout',     0.33)),
        'co2':      float(poids_raw.get('co2',      0.34)),
    }
    total = sum(poids.values()) or 1.0
    poids = {k: v / total for k, v in poids.items()}

    try:
        employes = [
            e.to_dict() for e in
            Employee.query.filter(
                Employee._id.in_(selected_emp_ids),
                Employee.lat != None
            ).all()
        ]
        dest_entity = Destination.query.filter_by(_id=dest_id).first()
        destination = dest_entity.to_dict() if dest_entity else None

        if not employes or not destination:
            return jsonify({'status': 'error', 'message': 'Données insuffisantes'}), 400

        result = generer_suggestions(employes, destination, capacite, poids=poids)
        return jsonify({'status': 'success', 'suggestions': result, 'count': len(result)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# OPTIMIZE STREAM (SSE)
@app.route('/api/optimize_stream', methods=['POST'])
def optimize_stream():
    import queue
    import threading
    from flask import Response

    data = request.json or {}
    selected_emp_ids = data.get('selected_employes_ids', [])
    dest_id = data.get('destination_id')
    capacite = _normaliser_capacite(data.get('capacite', 3))

    poids_raw = data.get('poids', {})
    poids = {
        'distance': float(poids_raw.get('distance', 0.33)),
        'cout':     float(poids_raw.get('cout',     0.33)),
        'co2':      float(poids_raw.get('co2',      0.34)),
    }
    total = sum(poids.values()) or 1.0
    poids = {k: v / total for k, v in poids.items()}

    employes = [
        e.to_dict() for e in
        Employee.query.filter(
            Employee._id.in_(selected_emp_ids),
            Employee.lat != None
        ).all()
    ]
    dest_entity = Destination.query.filter_by(_id=dest_id).first()
    destination = dest_entity.to_dict() if dest_entity else None

    if not employes or not destination:
        def error_gen():
            yield f"data: {json.dumps({'step': 'error', 'message': 'Données insuffisantes'})}\n\n"
        return Response(error_gen(), mimetype='text/event-stream')

    q = queue.SimpleQueue()

    def progress_callback(step, pct, msg):
        q.put({"step": step, "pct": pct, "msg": msg})

    def worker():
        try:
            result = generer_suggestions(employes, destination, capacite, poids=poids, progress_callback=progress_callback)
            q.put({"step": "done", "pct": 100, "result": {"status": "success", "suggestions": result, "count": len(result)}})
        except Exception as e:
            import traceback; traceback.print_exc()
            q.put({"step": "error", "message": str(e)})
        finally:
            q.put(None)

    threading.Thread(target=worker).start()

    def generate():
        while True:
            msg = q.get()
            if msg is None:
                break
            yield f"data: {json.dumps(msg)}\n\n"
            if msg.get("step") in ["done", "error"]:
                break

    return Response(generate(), mimetype='text/event-stream')


# OPTIMIZE MANUAL
@app.route('/api/optimize_manual', methods=['POST'])
def optimize_manual():
    data = request.json or {}
    manual_cluster_ids = data.get('manual_cluster_ids', [])
    dest_id = data.get('destination_id')
    capacite = _normaliser_capacite(data.get('capacite', 3))

    poids_raw = data.get('poids', {})
    poids = {
        'distance': float(poids_raw.get('distance', 0.33)),
        'cout':     float(poids_raw.get('cout',     0.33)),
        'co2':      float(poids_raw.get('co2',      0.34)),
    }
    total = sum(poids.values()) or 1.0
    poids = {k: v / total for k, v in poids.items()}

    try:
        from moovly_system import generer_suggestions_manual

        employes = [e.to_dict() for e in Employee.query.filter(Employee.lat != None).all()]
        dest_entity = Destination.query.filter_by(_id=dest_id).first()
        destination = dest_entity.to_dict() if dest_entity else None

        if not manual_cluster_ids or not destination:
            return jsonify({'status': 'error', 'message': 'Données insuffisantes'}), 400

        result = generer_suggestions_manual(manual_cluster_ids, destination, employes, capacite, poids=poids)
        return jsonify({'status': 'success', 'suggestions': result, 'count': len(result)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500



# COMPARATEUR DE SCÉNARIOS
@app.route('/api/compare_scenarios', methods=['POST'])
def compare_scenarios():
    try:
        data = request.get_json(force=True, silent=True) or {}
        selected_emp_ids = data.get('selected_employes_ids', [])
        dest_id          = data.get('destination_id')
        capacites        = data.get('capacites', [2, 3, 4])
        current_capacity = _normaliser_capacite(data.get('current_capacity', 3))

        poids_raw = data.get('poids', {})
        poids = {
            'distance': float(poids_raw.get('distance', 0.33)),
            'cout':     float(poids_raw.get('cout',     0.33)),
            'co2':      float(poids_raw.get('co2',      0.34)),
        }
        total = sum(poids.values()) or 1.0
        poids = {k: v / total for k, v in poids.items()}

        employes = [
            e.to_dict() for e in
            Employee.query.filter(
                Employee._id.in_(selected_emp_ids),
                Employee.lat != None
            ).all()
        ]

        if not employes:
            employes_raw = data.get('employes_data', [])
            employes = [e for e in employes_raw if e.get('lat') is not None]

        dest_entity = Destination.query.filter_by(_id=dest_id).first()
        destination = dest_entity.to_dict() if dest_entity else None

        if not destination:
            destination = data.get('destination_data')

        if not employes:
            return jsonify({'status': 'error', 'message': 'Aucun employé trouvé'}), 400
        if not destination:
            return jsonify({'status': 'error', 'message': 'Destination introuvable'}), 400

        from concurrent.futures import ThreadPoolExecutor

        def process_capacity(cap):
            cap = int(cap)
            suggestions = generer_suggestions(employes, destination, cap, poids=poids)
            if not suggestions:
                return None
            best = next((s for s in suggestions if s.get('is_best')), suggestions[0])
            rse  = best.get('rse_metrics', {})
            return {
                'capacite':          cap,
                'nb_vehicules':      len(best['routes']),
                'distance_km':       best['distance_km'],
                'duree_min':         best['duree_min'],
                'tarif_tnd':         best['tarif']['final'],
                'co2_kg':            rse.get('co2_scenario_moovly', 0),
                'co2_saved_kg':      rse.get('co2_saved_kg', 0),
                'cost_saved_tnd':    0.0,
                'fill_rate_percent': rse.get('fill_rate_percent', 0),
            }

        scenarios = []
        with ThreadPoolExecutor(max_workers=len(capacites)) as executor:
            results = list(executor.map(process_capacity, capacites))

        scenarios = [s for s in results if s is not None]

        baseline_scenario = next((s for s in scenarios if s['capacite'] == current_capacity), None)

        if not baseline_scenario:
            baseline_suggestions = generer_suggestions(employes, destination, current_capacity, poids=poids)
            if baseline_suggestions:
                b_best = next((s for s in baseline_suggestions if s.get('is_best')), baseline_suggestions[0])
                baseline_cost = b_best['tarif']['final']
            else:
                baseline_cost = None
        else:
            baseline_cost = baseline_scenario['tarif_tnd']

        for sc in scenarios:
            if baseline_cost is not None:
                sc['cost_saved_tnd'] = baseline_cost - sc['tarif_tnd']
            else:
                sc['cost_saved_tnd'] = 0.0

        return jsonify({'status': 'success', 'scenarios': scenarios, 'nb_employes': len(employes)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500



# EXPORT EXCEL
@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    import pandas as pd
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data   = request.json or {}
    result = data.get('result', {})

    C = {
        'dark':    "FF0F172A", 'dark2':  "FF1E293B", 'slate':  "FF334155",
        'indigo':  "FF6366F1", 'green':  "FF10B981", 'green_d':"FF065F46",
        'green_l': "FFD1FAE5",'amber':  "FFF59E0B", 'red':    "FFEF4444",
        'white':   "FFFFFFFF", 'light':  "FFF1F5F9", 'muted':  "FF94A3B8",
        'pale_v':  ["FFE0E7FF","FFFCE7F3","FFD1FAE5","FFFEF3C7",
                    "FFE0F2FE","FFEEF2FF","FFF0FDF4","FFFDF4FF"],
    }

    def fill(c):   return PatternFill("solid", fgColor=c)
    def fw(s=11):  return Font(bold=True, color=C['white'], size=s, name="Calibri")
    def fd(s=11):  return Font(bold=True, color=C['dark2'], size=s, name="Calibri")
    def fn(s=10, c=None): return Font(size=s, color=c or C['dark2'], name="Calibri")
    def ac():      return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def al():      return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def brd():
        s = Side(style="thin", color="FFE2E8F0")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_cell(ws, row, col, val, fnt=None, fll=None, aln=None, bor=None, h=None):
        c = ws.cell(row=row, column=col, value=val)
        if fnt: c.font      = fnt
        if fll: c.fill      = fll
        if aln: c.alignment = aln
        if bor: c.border    = bor
        if h:   ws.row_dimensions[row].height = h
        return c

    def auto_w(ws, mn=10, mx=45):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=mn)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+4, mn), mx)

    try:
        routes      = result.get('routes', [])
        dist_jour   = result.get('distance_km', 0)
        duree_jour  = result.get('duree_min', 0)
        tarif_jour  = result.get('tarif', {}).get('final', 0)
        rse         = result.get('rse_metrics', {})
        methode     = result.get('methode', 'Optimisation Moovly')
        destination = result.get('destination', {})
        nb_emp      = sum(len(r.get('ordre', [])) for r in routes)
        nb_veh      = len(routes)
        now_str     = datetime.now().strftime("%d/%m/%Y %H:%M")

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            wb = writer.book

            ws1 = wb.create_sheet("📊 Synthèse")
            ws1.merge_cells("A1:F1")
            set_cell(ws1,1,1,"MOOVLY — Rapport d'Optimisation de Flotte",
                     fw(15), fill(C['dark']), ac(), h=42)
            ws1.merge_cells("A2:F2")
            set_cell(ws1,2,1,f"Généré le {now_str}  •  Algorithme : {methode}  •  Destination : {destination.get('nom','—')}",
                     fn(9, C['muted']), fill(C['dark2']), ac(), h=20)
            ws1.row_dimensions[3].height = 6

            kpi_hdrs = ["Indicateur", "Valeur", "Unité", "× 5 jours", "× 20 jours", "Remarque"]
            kpi_data = [
                ["Véhicules mobilisés",      nb_veh,                                 "véhicules",nb_veh,                             nb_veh,                                   "Flotte du jour"],
                ["Employés transportés",     nb_emp,                                 "employés", nb_emp,                             nb_emp,                                   ""],
                ["Distance totale",           round(dist_jour,2),                    "km",       round(dist_jour*5,2),               round(dist_jour*20,2),                    "Routes OSRM réelles"],
                ["Durée totale",              round(duree_jour,1),                   "min",      round(duree_jour*5,1),              round(duree_jour*20,1),                   ""],
                ["Coût flotte (TND)",         round(tarif_jour,2),                   "TND",      round(tarif_jour*5,2),              round(tarif_jour*20,2),                   "Tarif officiel tunisien"],
                ["CO₂ émis Moovly (kg)",      rse.get('co2_scenario_moovly',0),      "kg",       round(rse.get('co2_scenario_moovly',0)*5,2), round(rse.get('co2_scenario_moovly',0)*20,2), "0,12 kg/km"],
                ["CO₂ évité (kg)",            rse.get('co2_saved_kg',0),             "kg",       round(rse.get('co2_saved_kg',0)*5,2),round(rse.get('co2_saved_kg',0)*20,2),  "vs voitures individuelles"],
                ["Économie financière (TND)", rse.get('cost_saved_tnd',0),           "TND",      round(rse.get('cost_saved_tnd',0)*5,2),round(rse.get('cost_saved_tnd',0)*20,2),"vs coût individuel"],
                ["Taux de remplissage",        rse.get('fill_rate_percent',0),        "%",        rse.get('fill_rate_percent',0),     rse.get('fill_rate_percent',0),           "Capacité utilisée"],
            ]

            for ci, h in enumerate(kpi_hdrs, 1):
                set_cell(ws1,4,ci,h, fw(10), fill(C['indigo']), ac(), brd(), h=28)

            for ri, row in enumerate(kpi_data, 5):
                bg = C['light'] if ri % 2 == 0 else C['white']
                for ci, val in enumerate(row, 1):
                    fnt = fd(11) if ci==1 else fn(11)
                    aln = al() if ci==1 else ac()
                    set_cell(ws1,ri,ci,val, fnt, fill(bg), aln, brd(), h=22)

            for ci in range(1,7):
                c = ws1.cell(row=12, column=ci)
                c.fill = fill(C['green_l'])
                c.font = Font(bold=True, color=C['green_d'], size=11, name="Calibri")

            auto_w(ws1); ws1.freeze_panes = "A5"

            ws2 = wb.create_sheet("🚗 Véhicules & Passagers")
            ws2.merge_cells("A1:G1")
            set_cell(ws2,1,1,"Détail des véhicules — Ordre de ramassage",fw(13),fill(C['dark']),ac(),h=32)
            h2 = ["Véhicule","# Arrêt","Passager","Résidence","Dist. cumulée (km)","Durée cumulée (min)","Tarif (TND)"]
            for ci,h in enumerate(h2,1):
                set_cell(ws2,2,ci,h, fw(10), fill(C['slate']), ac(), brd(), h=26)

            cur = 3
            for vi, route in enumerate(routes):
                vf = C['pale_v'][vi % len(C['pale_v'])]
                ordre = route.get('ordre',[])
                segs  = route.get('segments',[])
                d_cum = 0; t_cum = 0
                for si, emp in enumerate(ordre):
                    if si < len(segs):
                        d_cum += segs[si].get('distance_km',0)
                        t_cum += segs[si].get('duree_min',0)
                    row_v = [route.get('vehicule_id',''), si+1, emp.get('nom',''), emp.get('residence',''),
                             round(d_cum,2), round(t_cum,1),
                             route['tarif']['final'] if si==len(ordre)-1 else '']
                    for ci,val in enumerate(row_v,1):
                        set_cell(ws2,cur,ci,val,
                                 Font(bold=(ci==1),size=10,name="Calibri",color=C['dark2']),
                                 fill(vf), ac() if ci!=4 else al(), brd(), h=20)
                    cur += 1
                dest_row = [route.get('vehicule_id',''),"🏁",destination.get('nom',''),"— DESTINATION —",
                            round(route.get('distance_km',0),2),round(route.get('duree_min',0),1),
                            round(route['tarif']['final'],2)]
                for ci,val in enumerate(dest_row,1):
                    set_cell(ws2,cur,ci,val, fw(10), fill(C['dark2']), ac(), brd(), h=22)
                cur += 2
            auto_w(ws2); ws2.freeze_panes = "A3"

            ws3 = wb.create_sheet("🗺️ Segments")
            ws3.merge_cells("A1:F1")
            set_cell(ws3,1,1,"Segments de trajet détaillés par véhicule",fw(13),fill(C['dark']),ac(),h=32)
            h3 = ["Véhicule","Segment #","De","Vers","Distance (km)","Durée (min)"]
            for ci,h in enumerate(h3,1):
                set_cell(ws3,2,ci,h, fw(10), fill(C['slate']), ac(), brd(), h=26)
            sr = 3
            for vi, route in enumerate(routes):
                vf = C['pale_v'][vi % len(C['pale_v'])]
                for si, seg in enumerate(route.get('segments',[]),1):
                    bg = vf if sr%2==0 else C['white']
                    row_s = [route.get('vehicule_id',''),si,seg.get('from',''),seg.get('to',''),
                             seg.get('distance_km',0),seg.get('duree_min',0)]
                    for ci,val in enumerate(row_s,1):
                        set_cell(ws3,sr,ci,val, fn(10), fill(bg), ac() if ci not in[3,4] else al(), brd(), h=20)
                    sr+=1
                total_s = [route.get('vehicule_id',''),"TOTAL","—","—",
                           round(route.get('distance_km',0),2),round(route.get('duree_min',0),1)]
                for ci,val in enumerate(total_s,1):
                    set_cell(ws3,sr,ci,val, fw(10), fill(C['indigo']), ac(), brd(), h=22)
                sr+=2
            auto_w(ws3); ws3.freeze_panes = "A3"

            ws4 = wb.create_sheet("🌿 Impact RSE")
            ws4.merge_cells("A1:D1")
            set_cell(ws4,1,1,"Impact Environnemental & Économique",
                     Font(bold=True,size=15,color=C['white'],name="Calibri"),
                     fill("FF064E3B"), ac(), h=40)
            ws4.merge_cells("A2:D2")
            set_cell(ws4,2,1,"Flotte Moovly vs Trajets Individuels en voiture",
                     fn(10,C['muted']), fill("FF065F46"), ac(), h=20)

            rse_hdrs = ["Métrique","Voitures Individuelles","Flotte Moovly","Gain / Économie"]
            hdr_clrs = [C['slate'],"FFDC2626","FF059669","FF065F46"]
            for ci,h in enumerate(rse_hdrs,1):
                set_cell(ws4,4,ci,h, fw(11), fill(hdr_clrs[ci-1]), ac(), brd(), h=28)

            rse_rows = [
                ["CO₂ émis (kg)",           rse.get('co2_scenario_perso',0), rse.get('co2_scenario_moovly',0), rse.get('co2_saved_kg',0)],
                ["Coût transport (TND)",     rse.get('cost_scenario_perso',0), round(rse.get('cost_scenario_perso',0)-rse.get('cost_saved_tnd',0),2), rse.get('cost_saved_tnd',0)],
                ["Distance totale (km)",     round(dist_jour+rse.get('distance_saved_km',0),2), dist_jour, rse.get('distance_saved_km',0)],
                ["Taux remplissage (%)",     "1 pers/véhicule", rse.get('fill_rate_percent',0), "—"],
                ["Nb. véhicules",            nb_emp, nb_veh, nb_emp - nb_veh],
            ]
            for ri, row in enumerate(rse_rows, 5):
                bg = C['light'] if ri%2==0 else C['white']
                for ci,val in enumerate(row,1):
                    fnt = Font(bold=(ci in[1,4]),size=11,
                               color=C['green_d'] if ci==4 else C['dark2'],name="Calibri")
                    fll = fill(C['green_l']) if ci==4 else fill(bg)
                    set_cell(ws4,ri,ci,val, fnt, fll, ac() if ci!=1 else al(), brd(), h=24)

            ws4.row_dimensions[10].height = 10
            ws4.merge_cells("A11:D11")
            set_cell(ws4,11,1,"Projection annuelle (252 jours ouvrés)",fw(11),fill(C['indigo']),ac(),h=28)

 
            for ri,(label,val) in enumerate(annual, 12):
                set_cell(ws4,ri,1,label, fd(11), fill(C['light']), al(), brd(), h=24)
                set_cell(ws4,ri,2,val, Font(bold=True,size=12,color=C['green_d'],name="Calibri"),
                         fill(C['green_l']), ac(), brd())

            auto_w(ws4)

            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        output.seek(0)
        fname = f"moovly_rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# SQLITE

@app.route('/api/save_historique', methods=['POST'])
def save_historique():
    try:
        data = request.json
        if data.get('_clear'):
            OptimizationHistory.query.delete()
            db.session.commit()
            return jsonify({'status': 'success'})

        snap = OptimizationHistory(
            nb_employes     = data.get("nb_employes", 0),
            nb_vehicules    = data.get("nb_vehicules", 0),
            capacite        = str(data.get("capacite", "")),
            distance_km     = data.get("distance_km", 0),
            duree_min       = data.get("duree_min", 0),
            tarif_tnd       = data.get("tarif_tnd", 0),
            co2_kg          = data.get("co2_kg", 0),
            co2_saved_kg    = data.get("co2_saved_kg", 0),
            cost_saved_tnd  = data.get("cost_saved_tnd", 0),
            methode         = data.get("methode", ""),
            destination_nom = data.get("destination_nom", "")
        )
        db.session.add(snap)
        db.session.commit()
        return jsonify({'status': 'success', 'id': snap.snapshot_id})
    except Exception as e:
        db.session.rollback()
        print(f"Error saving historique: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/get_historique', methods=['GET'])
def get_historique():
    try:
        rows = OptimizationHistory.query.order_by(OptimizationHistory.date.asc()).all()
        hist = [r.to_dict() for r in rows]
        return jsonify({'status': 'success', 'historique': hist, 'count': len(hist)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500



# PARAMÈTRES
import moovly_system as ms

@app.route('/api/get_params', methods=['GET'])
def get_params():
    return jsonify({
        'status': 'success',
        'tarifs':     ms.TARIFS.copy(),
        'params_rse': ms.PARAMS_RSE.copy(),
        'params_algo':ms.PARAMS_ALGO.copy(),
    })

@app.route('/api/update_params', methods=['POST'])
def update_params():
    try:
        data = request.get_json()

        if 'tarifs' in data:
            t = data['tarifs']
            ms.TARIFS['prise_en_charge'] = float(t.get('prise_en_charge', ms.TARIFS['prise_en_charge']))
            ms.TARIFS['prix_79m']        = float(t.get('prix_79m',        ms.TARIFS['prix_79m']))
            ms.TARIFS['prix_18s']        = float(t.get('prix_18s',        ms.TARIFS['prix_18s']))
            ms.TARIFS['coef_hp']         = float(t.get('coef_hp',         ms.TARIFS['coef_hp']))
            ms.TARIFS['coef_weekend']    = float(t.get('coef_weekend',    ms.TARIFS['coef_weekend']))
            ms.TARIFS['mode_actif']      = str(t.get('mode_actif',        ms.TARIFS['mode_actif']))

        if 'params_rse' in data:
            r = data['params_rse']
            ms.PARAMS_RSE['co2_kg_per_km']          = float(r.get('co2_kg_per_km',          ms.PARAMS_RSE['co2_kg_per_km']))
            ms.PARAMS_RSE['cout_individuel_per_km']  = float(r.get('cout_individuel_per_km', ms.PARAMS_RSE['cout_individuel_per_km']))
            if 'fleet_composition' in r:
                ms.PARAMS_RSE['fleet_composition'] = r['fleet_composition']

        if 'params_algo' in data:
            a = data['params_algo']
            ms.PARAMS_ALGO['max_cluster_distance_km'] = float(a.get('max_cluster_distance_km', ms.PARAMS_ALGO['max_cluster_distance_km']))
            ms.PARAMS_ALGO['ortools_time_limit_s']    = int(float(a.get('ortools_time_limit_s',    ms.PARAMS_ALGO['ortools_time_limit_s'])))
            ms.PARAMS_ALGO['direction_lambda']        = float(a.get('direction_lambda',        ms.PARAMS_ALGO['direction_lambda']))

        return jsonify({'status': 'success', 'message': 'Paramètres mis à jour.'})
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()}), 500


# MAIN
if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 MOOVLY FLEET OPTIMIZER — SQLite Edition")
    print("📍 http://127.0.0.1:5001")
    print("="*60 + "\n")
    app.run(debug=True, use_reloader=False, port=5001)